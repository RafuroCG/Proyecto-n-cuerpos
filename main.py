import random
import string
import tkinter as tk
import tkinter.colorchooser as colorchooser

import matplotlib.pyplot as plt
import matplotlib.animation as anim
import numpy as np
import openpyxl # Ejecutar en la terminal "python3 -m pip install openpyx"

from numpy.linalg import norm
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import ttk, filedialog, messagebox

from Body_file import Body

# Constante gravitacional en Unidades astronómicas, masas solares y años.
G = 4 * np.pi**2 

class SimulationParameters:
    '''
    Contiene los parámetros de la simulación

    Atributos:
        dt (float): El valor de diferencia entre dos tiempos consecutivos (delta t).
        user_dt (float): El valor de dt ingresado por el usuario.
        eps (float): El valor de épsilon para el algoritmo de corrección.
        correctAlg_enabled (bool): Estado del algormitmo de corrección (Activado o desactiado)
        animation_interval (int): milisegundos entre cada iteración de la animación.

    Métodos: N/A
    '''
    def __init__(self):
        self.user_dt = 0.001
        self.dt = self.user_dt
        self.eps = 5.0
        self.animation_interval = 50
        self.correctAlg_enabled = False


class ParticleManager:
    '''
    Contiene las funciones que manejarán las partículas y sus parámetros iniciales en la simulación.
    
    Atributos:
        bodies (list): La lista de las partículas en la simulación. Es una lista de tuplas cada una de las 
        cuales contiene el objeto, sus posiciones en x, sus posiciones en y, su gráfica (ax.plot) y sus
        velocidades a lo largo de la simulación.
        Xall (list): La lista de todas las posiciones en x de todas las partículas a lo largo de toda la 
        simulación.
        Yall (list): La lista de todas las posiciones en x de todas las partículas a lo largo de toda la 
        simulación.
        frame_data (list): Los datos del cuadro actual de la simulación.
        saved_filename (str): Nombre del archivo a guardar si el usuario desea guardar los datos simulados.

    Métodos:
        add_particle(masa, pos0, vel0, color, name="body"): Crea la partícula con los parámetros 
        iniciales (masa, posición y velocidad inicial), color y nombre indicados, la agrega a 
        la lista de partículas y crea su correspondiente curva en la gráfica.

        generate_random_particles(num_particles): Genera un número especificado de partículas aleatorias 
        haciendo uso del módulo random y funciones random_color() y random_name().

        clear_particles(): Borra todas las partículas de la lista, eliminándolas de la simulación y 
        borra los datos de la gráfica y de Xall y Yall.

        save_particle_data(): Maneja la opción de guardar la información actual de la simulación. Usa 
        la librería openpyxl para crear un archivo de tipo .xlsl (Excel) donde se va a guardar la 
        masa, posición, velocidad y color de cada partícula, si el usuario lo desea.

        save_frame_data():
    '''
    def __init__(self):
        self.bodies = []
        self.Xall = []
        self.Yall = []
        self.frame_data = []


    def add_particle(self, masa, pos0, vel0, color, name="body"):
        '''
        Crea la partícula con los parámetros iniciales (masa, posición y velocidad inicial), color 
        y nombre indicados, la agrega a la lista de partículas y crea su correspondiente curva 
        en la gráfica.

        :param masa: Masa de la partícula. Flotante.
        :param pos0: Posición inicial de la partícula. Tupla de R2.
        :param vel0: Velocidad inicial de la partícula. Tupla de R2.
        :param color: Representación hexadecimal del color de la partícula. Cadena de caracteres.
        :return: N/A.
        '''
        new_body = Body(masa, pos0, vel0, color, name) # Crea un cuerpo (objeto de la clase Body) con los 
        # parámetros indicados.
        line_obj, = ax.plot([], [], color=color, linestyle='-', linewidth=1) # Crea la gráfica del objeto.
        self.bodies.append((new_body, [new_body.pos[0]], [new_body.pos[1]], line_obj, [new_body.vel]))
        # Se agrega la información de la partícula a la lista de partículas 'bodies'.
    

    def generate_random_particles(self, num_particles):
        '''
        Genera un número especificado de partículas aleatorias haciendo uso del módulo random y 
        funciones random_color() y random_name().

        :param num_particles: número de partículas a generar. Entero positivo.
        :return: N/A.
        '''

        for k in range(num_particles):
            # Se inicializan los parámetros aleatoriamente para cada partícula a agregar.
            masa = random.uniform(1, 20)
            pos0 = (random.uniform(-10, 10), random.uniform(-10, 10))
            vel0 = (random.uniform(-10, 10), random.uniform(-10, 10))
            color = random_color()
            name = random_name()
            new_body = Body(masa, pos0, vel0, color, name)
            line_obj, = ax.plot([], [], color=color, linestyle='-', linewidth=1) # Se inicializa la gráfica 
            # de cada partícula.
            self.bodies.append((new_body, [new_body.pos[0]], [new_body.pos[1]], line_obj, [new_body.vel]))
            # Se ingresa la partícula a la lista.
    

    def clear_particles(self):
        '''
        Borra todas las partículas de la lista, eliminándolas de la simulación y borra los datos de la 
        gráfica y de Xall y Yall.
        '''
        self.bodies = []
        self.Xall.clear()
        self.Yall.clear()
        ax.clear()
        ax.grid(True) 
        fig_canvas.draw()


    def save_particle_data(self):
        '''
        Función encargada de crear un archivo .xlsx con los datos de las partículas en el tiempo actual e 
        crear una segunda hoja para posteriormente guardar la información de la simulación en cada tiempo 
        (frame).

        :return: N/A
        '''
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", \
                                                filetypes=[("Excel files", "*.xlsx")])
        
        if filename:
            # Si se ingresó un nombre, se crea un archivo .xlsx.
            wb = openpyxl.Workbook()
            ws = wb.active # Accede a la hoja 1 creada por defecto.
            ws.title = "Particle Data"
            headers = ["Name", "Mass", "Position X", "Position Y", "Velocity X", "Velocity Y", "Color"]
            ws.append(headers) # La primera fila tendrá los nombres de las variables a organizar.
            header_font = Font(bold=True) # Configura el estilo de las palabras del encabezado a negrilla.
            # Configura las líneas de las casillas del encabezado a delgadas.
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                        top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            
            # Configura las casillas de la primera fila con los atributos anteriores.
            for cell in ws[1]: 
                cell.font = header_font
                cell.border = border
                # Alínea los textos al centro de las casillas.
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Accede a las tuplas de 'bodies' para guardar la posición y velocidad actual de cada 
            # partícula en la hoja de cálculo (solo se usa el primer elemento, que corresponde al 
            # objeto de cada partícula).
            for body, _, _, _, _ in self.bodies:
                ws.append([
                    body.name,
                    body.m,
                    body.pos[0],
                    body.pos[1],
                    body.vel[0],
                    body.vel[1],
                    body.color
                ])  

            # Se configura cada fila después del encabezado (min_row = 2) de la misma forma que el 
            # encabezado sin las negrillas.
            for row in ws.iter_rows(min_row=2, max_col=len(headers), max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Itera sobre cada columna para ajustar su ancho dependiendo del elemento más largo.
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    # Se usa un "try: [...] \except: pass" para ignorar casillas vacías que aparecen cuando 
                    # dos o más partículas se crean en tiempos distintos.
                    try:
                        # 15 se escogió para no ajustar el ancho para flotantes tan largos por estética.
                        if len(str(cell.value)) > max_length and len(str(cell.value)) <= 15:
                            max_length = len(cell.value) # Se ajusta la longitud máxima a la máxima hasta 
                            # entonces.
                    except:
                            pass
                # Se le suman 2 espacios para no verse tan apiñados los valores.
                adjusted_width = max_length + 2 
                # Se ajusta el ancho de la columna actual.
                ws.column_dimensions[column_letter].width = adjusted_width
            # Se crea la segunda hoja donde se guardará la información por cada iteración (frame) de 
            # la simulación.
            ws_frames = wb.create_sheet(title="Frame Data")
            frame_headers = ["Frame"]
            # Se crea la lista del encabezado de dicha hoja donde van los nombres de las variables para 
            # cada masa
            for body in self.bodies:
                frame_headers.append(f"{body[0].name} PosX")
                frame_headers.append(f"{body[0].name} PosY")
                frame_headers.append(f"{body[0].name} VelX")
                frame_headers.append(f"{body[0].name} VelY")
            # Se escribe dicha lista como encabezado a la hoja.
            ws_frames.append(frame_headers)

            # Configura el encabezado de la misma forma que en la primera hoja.
            for cell in ws_frames[1]:
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajusta el ancho de las columnas de la misma forma que en la primera hoja.
            for column in ws_frames.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length and len(str(cell.value)) <= 15:
                            max_length = len(cell.value)
                    except:
                            pass
                adjusted_width = (max_length + 2)
                ws_frames.column_dimensions[column_letter].width = adjusted_width
            wb.save(filename)
            # Se guarda el nombre del archivo para actualizaciones en cada iteración; se usa en la 
            # próxima función.
            self.saved_filename = filename  


    def save_frame_data(self):
        '''
        Función encargada de guardar en la segunda hoja del archivo creado por la función anterior los 
        datos de la simulación en cada tiempo (frame).

        :return: None o N/A.
        '''
        # Se maneja el error de que el usuario no ingrese un nombre para el archivo.
        if not hasattr(self, 'saved_filename'):
            # El método "hasattr(obj, attr)" verifica que el objeto 'obj' tenga el atributo "attr"; en este
            # caso se verifica que en esta clase exista el atributo "saved_filename" creado al ingresar un
            # nombre para el archivo.
            messagebox.showerror("Error", "Primero guarda los datos de las partículas.")
            return # Para acabar la función retornando "None" en caso de no haber ingresado un nombre.
    
        # Se vuelve a abrir el libro de excel.
        wb = openpyxl.load_workbook(self.saved_filename)
        # Se accede a la segunda hoja.
        ws_frames = wb["Frame Data"]
        # Se guarda cada lista de "frame_data" de cada tiempo (frame) en cada fila de la segunda hoja. 
        for frame in self.frame_data:
            ws_frames.append(frame)
        
        # Guarda el archivo y avisa al usuario que ha sido guardado exitosamente.
        wb.save(self.saved_filename)
        messagebox.showinfo("Datos Guardados", f"Datos guardados en {self.saved_filename}")
    
    
# Este objeto contendrá los parámetros de la simulación.
simulation_params = SimulationParameters()
# Este objeto será el responsable de gestionar las partículas.
particle_manager = ParticleManager()


def random_color():
    '''
    Escoge un color al azar usando el método "randint" de la librería "random" hallando una 
    representación hexadecimal con números aleatorios.

    :return: Cadena de caracteres de representación hexadecimal del color.
    '''
    return '#%06x' % random.randint(0, 0xFFFFFF)# Genera un número de 6 dígitos (06) en hexadecimal (x).


def random_name(length=5):
    '''
    Arroja un nombre aleatorio usando el método "choices" de la librería "random" para escoger al azar 
    entre una letra mayúscula y un número de 0 a 9 5 veces para un nombre aleatorio de 5 caracteres. 
    Ejemplos: 6ALVF, 1O057, KFQB5.

    :length: Longitud del nombre inicializada a 5 caracteres. Número entero positivo.
    :return: Nombre aleatorio de 5 caracteres.
    '''
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=length))


# Dos funciones similares para verificar los valores ingresados de dt y epsilon
def verify_dt_input(entry, label):
    '''
    Función encargada de comprobar el valor ingresado de dt para ser asignado.

    :label: La etiqueta que muestra el valor actual de dt.
    :return: N/A.
    '''
    # Se comprueba que el usuario haya ingresado un valor numérico, si no, arroja un error que muestra
    # una advertencia.
    try:
        dt_value = float(entry.get()) # Intenta convertir a flotante.
        # Si el valor ingresado es menor que cero, arroja el error.
        if dt_value <= 0:
            raise ValueError("Time step must be positive.")
        # Si no se arrojó error, se asigna el valor al parámetro dt del objeto de la clase que contiene los
        # parámetros de la simulación.
        simulation_params.user_dt = dt_value
        # Se cambia la etiqueta para mostrar el nuevo valor de dt.
        label.config(text=str(dt_value))
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid number for time step (dt)")


def verify_eps_input(entry, label):
    '''
    Función encargada de comprobar el valor de "epsilon" para ser asignado. "epsilon" es el valor máximo 
    para la diferencia entre dos velocidades consecutivas que se usa en el algoritmo de corrección.

    :label: La etiqueta que muestra el valor actual de epsilon.
    :return: N/A
    '''
    # Refiérase a los comentarios de la función "verify_dt_input" puesto que sigue la misma lógica.
    try:
        eps_value = float(entry.get())
        if eps_value <= 0:
            raise ValueError("Epsilon must be positive.")
        simulation_params.eps = eps_value
        label.config(text=str(eps_value))
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid number for epsilon")


def add_particle():
    '''
    Función a ejecutar al presionar el botón que agrega partículas no aleatorias. Recibe todos los 
    parámetros ingresados y agrega la partícula con dichos parámetros usando la función anterior.

    :return: None o N/A.
    '''
    # Se reciben los parámetros.
    masa = masa_entry.get()
    posx = posx_entry.get()
    posy = posy_entry.get()
    velx = velx_entry.get()
    vely = vely_entry.get()
    color = color_entry.get()
    name = name_entry.get()
    # Si no ingresó alguno de los parámetros se advierte al usuario y finaliza la función.
    if not masa or not posx or not posy or not velx or not vely or not color or not name:
        messagebox.showwarning("Incomplete Data", "Please fill in all fields before adding a particle.")
        return
    # Si no se puede convertir alguno de los valores numéricos a como tales, se advierte al usuario y 
    # finaliza la función.
    try:
        masa = float(masa)
        pos0 = (float(posx), float(posy))
        vel0 = (float(velx), float(vely))
    except ValueError:
        messagebox.showerror("Invalid Input", "Mass, Position, and Velocity must be numeric values.")
        return
    # Agrega la partícula haciendo uso de la función anterior.
    particle_manager.add_particle(masa, pos0, vel0, color, name)


def generate_random_particles():
    '''
    Se ejecuta al presionar el botón de agregar partículas aleatorias, verificando la entrada del número
    deseado y, si no recibe errores, ejecutando la función generate_random_particles de la clase 
    "ParticleManager".

    :return: None o N/A.
    '''
    # Recibe el valor de la entrada del usuario.
    num_particles = num_particles_entry.get()
    # Verifica que haya un valor en la entrada al presionar el botón.
    if not num_particles:
        messagebox.showwarning("Input Required", "Please enter a number of particles.")
        return
    # Se verifica que el valor sea un entero positivo, si no, advierte al usuario.
    try:
        num_particles = int(num_particles)
        if num_particles <= 0:
            raise ValueError("Number of particles must be a positive integer.")
        particle_manager.generate_random_particles(num_particles)
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid integer for the number of particles.")
        return


def animate(frame):
    '''
    Función a ejecutar en cada iteración de la animación (parámetro del FuncAnimation). Actualiza la 
    lista de partículas del "particle_manager" con los datos de la iteración y ajusta la gráfica.

    :frame: número de la iteración actual.
    :return: None o N/A. 
    '''
    if len(particle_manager.bodies) == 0:
        return
    # "frame_data" es una lista que contiene la información de la iteración actual para ser guardada
    # en el archivo si el usuario lo desea.
    frame_data = [frame] 
    # Se accede a la lista de cuerpos de la simulación para ser actualizada.
    for body, x_data, y_data, line_obj, vel_data in particle_manager.bodies:

        # Pasos de la integración "leapfrog": Se actualiza la velocidad a la mitad del intervalo y 
        # se calcula la posición al final con dicha velocidad.
        old_vel = body.vel.copy()
        body.update_vel([b[0] for b in particle_manager.bodies], simulation_params.dt)
        body.update_pos(simulation_params.dt)
        new_vel = body.vel.copy()

        # Algoritmo de corrección  del delta t; si la magnitud de la diferencia de dos velocidades
        # es mayor a épsilon, el dt se corrige para que, dependiendo de la aceleración sufrida por 
        # el objeto, el cambio en la rapidez no sea mayor a este valor. Esto permite que, una vez se 
        # cumpla la desigualdad (se supere el épsilon), para la próxima iteración se corriga el dt.
        if simulation_params.correctAlg_enabled and norm(old_vel - new_vel) > simulation_params.eps:
            simulation_params.dt = simulation_params.eps / norm(body.r_accel)

        # Se agregan las posiciones y velocidades actuales de la partícula a sus respectivas listas
        # en la lista de partículas y a la lista de todas las posiciones en X y en Y.
        x_data.append(body.pos[0])
        y_data.append(body.pos[1])
        vel_data.append(body.vel)
        particle_manager.Xall.append(body.pos[0])
        particle_manager.Yall.append(body.pos[1])

        # Se actualiza la gráfica de la partícula con los datos actuales de posición.
        line_obj.set_data(x_data, y_data)
        
        # A frame_data se le agregan los datos actuales.
        frame_data.append(body.pos[0])
        frame_data.append(body.pos[1])
        frame_data.append(body.vel[0])
        frame_data.append(body.vel[1])
    
    # Revisa que el usuario todavía desee corregir. Si no lo desea, devuelve el dt al ingresado.
    if not simulation_params.correctAlg_enabled:
        simulation_params.dt = simulation_params.user_dt
    
    # Agrega los datos de la iteración a la lista de los datos de la simulación.
    particle_manager.frame_data.append(frame_data)
    
    # Estos "deltas" se usan para dejar un espaciado, que se escogió como 1/10 del rango de posiciones
    # en su respectivo eje, entre la partícula más lejana en cada eje y el borde de la gráfica. 
    axdelta = (max(particle_manager.Xall) - min(particle_manager.Xall))/10
    aydelta = (max(particle_manager.Yall) - min(particle_manager.Yall))/10
    # Se actualizan los límites de la gráfica para mostrar todas las trayectorias de las partículas:
    # si las partículas se acercan, la gráfica se mantiene; si las partículas se alejan, la gráfica 
    # se expande.
    ax.set_xlim(min(particle_manager.Xall) - axdelta, max(particle_manager.Xall) + axdelta)
    ax.set_ylim(min(particle_manager.Yall) - aydelta, max(particle_manager.Yall) + aydelta)
    print("\n")
    ax.grid(True)
    # Dibuja la nueva gráfica en el lienzo de tkinter.
    fig_canvas.draw()


animation = None
# En las siguientes tres funciones se usó "animation" como variable global para que su valor externo
# pudiese ser cambiado.
def start_animation(canvas):
    '''A ejecutar al presionar el botón que comienza la animación.'''
    global animation
    # Si no existía animación, se asigna a la variable la animación de "FuncAnimation"
    if animation is None:
        animation = anim.FuncAnimation(fig, animate, \
                                       interval=simulation_params.animation_interval, blit=False)
    canvas.draw()


def pause_animation():
    '''Pausa la animación si existe. Se ejecuta al presionar su respectivo botón.'''
    global animation
    if animation is not None:
        animation.event_source.stop()


def resume_animation():
    '''Continúa la animación si está pausada. Se ejcuta al presionar su respectivo botón.'''
    global animation
    if animation is not None:
        animation.event_source.start()


def close_window():
    '''Cierra la ventana de Tkinter.'''
    window.destroy()


def toggle_correct_alg(var):
    '''Activa el algoritmo de corrección al ser presionado su respectivo botón.'''
    print("***")
    simulation_params.correctAlg_enabled = var.get()


def choose_color():
    '''
    Permite al usuario escoger el color a partir de un cuadro dinámico que aparece al presionar 
    su botón.
    '''
    # Crea el cuadro dinámico de colores.
    color_code = colorchooser.askcolor(title="Choose color")[1]
    if color_code:
        # A la entrada le borra el valor ingresado anteriormente y lo cambia por el asignado por el cuadro.
        color_entry.delete(0, tk.END)
        color_entry.insert(0, color_code)

# Se crea la ventana de Tkinter.
window = tk.Tk()
window.title("Particle Simulation")

# Se crea la gráfica y se inicializa la subgráfica (subplot) en la cual se dibujará.
fig, ax = plt.subplots()
ax.set_xlim(0, 15)
ax.set_ylim(0, 15)
ax.grid(True)

# Se crea un cuadro (frame) de tkinter donde estarán colocados todos los controles de la simulación.
controls_frame = ttk.Frame(window)
controls_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="NSEW")

# A continuación se crean las etiquetas, entradas y botones (si aplican) de cada control:
# Control de dt:
ttk.Label(controls_frame, text="Time Step (dt)").grid(row=0, column=0, padx=10, pady=10)# Etiqueta
dt_input = ttk.Entry(controls_frame)# Entrada
dt_input.grid(row=0, column=1, padx=10, pady=10)
dt_button = ttk.Button(controls_frame, text="Assign dt", \
                       command=lambda: verify_dt_input(dt_input, dt_label))# Botón
dt_button.grid(row=0, column=2, padx=10, pady=10)
dt_label = ttk.Label(controls_frame, text=str(simulation_params.user_dt))# Etiqueta del valor.
dt_label.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

# Control de epsilon:
ttk.Label(controls_frame, text="Epsilon").grid(row=2, column=0, padx=10, pady=10)
eps_input = ttk.Entry(controls_frame)
eps_input.grid(row=2, column=1, padx=10, pady=10)
eps_button = ttk.Button(controls_frame, text="Assign Epsilon", \
                        command=lambda: verify_eps_input(eps_input, eps_label))
eps_button.grid(row=2, column=2, padx=10, pady=10)
eps_label = ttk.Label(controls_frame, text=str(simulation_params.eps))
eps_label.grid(row=3, column=0, columnspan=3, padx=10, pady=10)

# Control de selección de algoritmo de corrección:
correctAlg_checkbutton_var = tk.BooleanVar()# Variable de tipo booleana (True o False)
correctAlg_checkbutton = ttk.Checkbutton(
    controls_frame, text="Enable Correction Algorithm", variable=correctAlg_checkbutton_var,
    command=lambda: toggle_correct_alg(correctAlg_checkbutton_var)
)# Botón de selección
correctAlg_checkbutton.grid(row=4, column=0, columnspan=3, padx=10, pady=10)

# Para crear una partícula:
# Masa:
ttk.Label(controls_frame, text="Mass").grid(row=5, column=0, padx=10, pady=10)
masa_entry = ttk.Entry(controls_frame)
masa_entry.grid(row=5, column=1, padx=10, pady=10)

# Posición inicial en x:
ttk.Label(controls_frame, text="Position X").grid(row=6, column=0, padx=10, pady=10)
posx_entry = ttk.Entry(controls_frame)
posx_entry.grid(row=6, column=1, padx=10, pady=10)

# Posición inicial en y:
ttk.Label(controls_frame, text="Position Y").grid(row=7, column=0, padx=10, pady=10)
posy_entry = ttk.Entry(controls_frame)
posy_entry.grid(row=7, column=1, padx=10, pady=10)

# Velocidad inicial en x:
ttk.Label(controls_frame, text="Velocity X").grid(row=8, column=0, padx=10, pady=10)
velx_entry = ttk.Entry(controls_frame)
velx_entry.grid(row=8, column=1, padx=10, pady=10)

# Velocidad inicial en y:
ttk.Label(controls_frame, text="Velocity Y").grid(row=9, column=0, padx=10, pady=10)
vely_entry = ttk.Entry(controls_frame)
vely_entry.grid(row=9, column=1, padx=10, pady=10)

# Color:
ttk.Label(controls_frame, text="Color").grid(row=10, column=0, padx=10, pady=10)
color_entry = ttk.Entry(controls_frame)
color_entry.grid(row=10, column=1, padx=10, pady=10)
color_button = ttk.Button(controls_frame, text="Choose Color", command=choose_color)
color_button.grid(row=10, column=2, padx=10, pady=10)

# Nombre:
ttk.Label(controls_frame, text="Name").grid(row=11, column=0, padx=10, pady=10)
name_entry = ttk.Entry(controls_frame)
name_entry.grid(row=11, column=1, padx=10, pady=10)

# Botón para agregar la partícula con los parámetros anteriores:
add_button = ttk.Button(
    controls_frame, text="Add Particle", command=add_particle
)
add_button.grid(row=12, column=0, columnspan=3, padx=10, pady=10)

# Para generar partículas aleatorias:
ttk.Label(controls_frame, text="Number of particles").grid(row=13, column=0, padx=10, pady=10)
num_particles_entry = ttk.Entry(controls_frame)
num_particles_entry.grid(row=13, column=1, padx=10, pady=10)
generate_button = ttk.Button(
    controls_frame, text="Generate Random Particles", command=generate_random_particles
)
generate_button.grid(row=13, column=2, padx=10, pady=10)

# Para comenzar la animación:
begin_button = ttk.Button(
    controls_frame, text="Begin Animation", command=lambda: start_animation(fig_canvas)
)
begin_button.grid(row=0, column=3, padx=10, pady=10)

# Para pausar la animación:
pause_button = ttk.Button(
    controls_frame, text="Pause Animation", command=pause_animation
)
pause_button.grid(row=1, column=3, padx=10)

# Para continuar la animación:
resume_button = ttk.Button(
    controls_frame, text="Continue Animation", command=resume_animation
)
resume_button.grid(row=2, column=3, padx=10)

# Para borrar todas las partículas:
clear_button = ttk.Button(
    controls_frame, text="Clear Particles", command=particle_manager.clear_particles
)
clear_button.grid(row=3, column=3, columnspan=3, padx=10, pady=10)

# Para cerrar la ventana:
close_button = ttk.Button(
    controls_frame, text="Close", command=close_window
)
close_button.grid(row=5, column=3, padx=10)

# Para guardar la información de la simulación en un archivo:
save_button = ttk.Button(
    controls_frame, text="Save current data", \
    command=lambda: [particle_manager.save_particle_data(), particle_manager.save_frame_data()]
)
save_button.grid(row=4, column=3, padx=10)

# Se crea otro cuadro (frame) para la gráfica.
fig_frame = ttk.Frame(window)
fig_frame.grid(row = 0, column = 4) # Se coloca a la derecha de todos los otros botones.
fig_canvas = FigureCanvasTkAgg(fig, master=fig_frame)
fig_canvas.get_tk_widget().grid(row=0, column=1, sticky="NSEW")
window.grid_rowconfigure(0, weight=1)
window.grid_columnconfigure(1, weight=1)
window.mainloop()